# economic_sim_web.py
import os
from io import BytesIO
import pandas as pd
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple, Union,Iterable
from enum import Enum
from datetime import datetime, timedelta
import base64
from copy import deepcopy
from flask import send_file,flash,url_for,Flask, request, render_template_string, redirect,session

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    print("Warning: openpyxl package not found. Excel export functionality will be disabled.")
    print("To enable Excel export, please install openpyxl using: pip install openpyxl")
    EXCEL_AVAILABLE = False
# ======== 原有经济模拟核心逻辑 ========
# ... [保持原有的Enum定义、dataclass和核心类不变] ...
# （此处完整包含问题中提供的所有类定义：AgentType到EconomicSystem）
# Core Enums and Classes
class AgentType(Enum):
    BANK = "bank"
    COMPANY = "company"
    HOUSEHOLD = "household"
    TREASURY = "treasury"
    CENTRAL_BANK = "central_bank"
    OTHER = "other"

class EntryType(Enum):
    LOAN = "loan"
    DEPOSIT = "deposit"  # No maturity or settlement type
    PAYABLE = "payable"  # Always means of payment settlement
    BOND = "bond"
    DELIVERY_CLAIM = "delivery_claim"  # Always non-financial asset settlement
    NON_FINANCIAL = "non_financial"  # No maturity or settlement type
    DEFAULT = "default"  # Used when settlement fails

class MaturityType(Enum):
    ON_DEMAND = "on_demand"
    FIXED_DATE = "fixed_date"
    PERPETUAL = "perpetual"

class SettlementType(Enum):
    MEANS_OF_PAYMENT = "means_of_payment"  # Standard payment means (bank transfer, check, etc.)
    SECURITIES = "securities"  # Financial instruments
    NON_FINANCIAL_ASSET = "non_financial_asset"  # Physical assets
    SERVICES = "services"  # Services rendered
    CRYPTO = "crypto"  # Cryptocurrency
    NONE = "none"  # No settlement type

@dataclass
class SettlementDetails:
    type: SettlementType
    denomination: str  # Currency or unit of settlement

@dataclass
class BalanceSheetEntry:
    type: EntryType
    is_asset: bool  # True for assets, False for liabilities
    counterparty: Optional[str]  # Optional for non-financial entries
    amount: float
    denomination: str
    maturity_type: MaturityType
    maturity_date: Optional[datetime]  # Required for FIXED_DATE, None for others
    settlement_details: SettlementDetails
    name: Optional[str] = None  # For non-financial assets or special naming
    issuance_time: str = 't0'  # When the entry was created (t0, t1, t2)

    def matches(self, other: 'BalanceSheetEntry') -> bool:
        """Check if two entries match (used for removing entries)"""
        return (
            self.type == other.type and
            self.is_asset == other.is_asset and
            self.counterparty == other.counterparty and
            self.amount == other.amount and
            self.denomination == other.denomination and
            self.maturity_type == other.maturity_type and
            self.maturity_date == other.maturity_date and
            self.settlement_details.type == other.settlement_details.type and
            self.settlement_details.denomination == other.settlement_details.denomination and
            self.name == other.name and
            self.issuance_time == other.issuance_time
        )

    def __post_init__(self):
        if self.amount <= 0:
            raise ValueError("Amount must be positive")

        # Validate issuance time
        if self.issuance_time not in ['t0', 't1', 't2']:
            raise ValueError("Issuance time must be 't0', 't1', or 't2'")

        # Validate counterparty rules
        if self.type != EntryType.NON_FINANCIAL and not self.counterparty:
            raise ValueError("Counterparty is required for financial entries")
        if self.type == EntryType.NON_FINANCIAL and self.counterparty:
            raise ValueError("Non-financial entries cannot have a counterparty")

        # Validate name rules
        if self.type == EntryType.NON_FINANCIAL and not self.name:
            raise ValueError("Non-financial entries must have a name")

        # Validate payable rules
        if self.type == EntryType.PAYABLE and self.settlement_details.type != SettlementType.MEANS_OF_PAYMENT:
            raise ValueError("Payable entries must have means_of_payment settlement type")

class SettlementFailure(Exception):
    def __init__(self, agent_name: str, entry: BalanceSheetEntry, reason: str):
        self.agent_name = agent_name
        self.entry = entry
        self.reason = reason
        super().__init__(f"Settlement failure for {agent_name}: {reason}")

class Agent:
    def __init__(self, name: str, agent_type: AgentType):
        self.name = name
        self.type = agent_type
        self.assets: List[BalanceSheetEntry] = []
        self.liabilities: List[BalanceSheetEntry] = []
        self.status: str = "operating"  # operating or bankrupt
        self.creation_time: datetime = datetime.now()
        # Add settlement history
        self.settlement_history = {
            'as_asset_holder': [],  # Settlements where this agent was the creditor
            'as_liability_holder': []  # Settlements where this agent was the debtor
        }

    def add_asset(self, entry: BalanceSheetEntry):
        self.assets.append(entry)

    def add_liability(self, entry: BalanceSheetEntry):
        self.liabilities.append(entry)

    def remove_asset(self, entry: BalanceSheetEntry):
        self.assets = [e for e in self.assets if not e.matches(entry)]

    def remove_liability(self, entry: BalanceSheetEntry):
        self.liabilities = [e for e in self.liabilities if not e.matches(entry)]

    def get_balance_sheet(self) -> Dict:
        return {
            "assets": self.assets,
            "liabilities": self.liabilities
        }

    def get_total_assets(self) -> float:
        return sum(entry.amount for entry in self.assets)

    def get_total_liabilities(self) -> float:
        return sum(entry.amount for entry in self.liabilities)

    def get_net_worth(self) -> float:
        return self.get_total_assets() - self.get_total_liabilities()

    def get_type_specific_metrics(self) -> Dict:
        metrics = {
            "name": self.name,
            "type": self.type.value,
            "creation_time": self.creation_time,
            "status": self.status,
            "total_assets": self.get_total_assets(),
            "total_liabilities": self.get_total_liabilities(),
            "net_worth": self.get_net_worth()
        }

        if self.type == AgentType.BANK:
            metrics["capital_ratio"] = self.get_total_assets() / self.get_total_liabilities() if self.get_total_liabilities() > 0 else float('inf')
        elif self.type == AgentType.COMPANY:
            metrics["leverage_ratio"] = self.get_total_liabilities() / self.get_total_assets() if self.get_total_assets() > 0 else float('inf')
        elif self.type == AgentType.HOUSEHOLD:
            metrics["savings_rate"] = (self.get_total_assets() - self.get_total_liabilities()) / self.get_total_assets() if self.get_total_assets() > 0 else 0

        return metrics

    def record_settlement(self,
                         time_point: str,
                         original_entry: BalanceSheetEntry,
                         settlement_result: BalanceSheetEntry,
                         counterparty: str,
                         as_asset_holder: bool):
        """Record a settlement in the agent's history"""
        settlement_record = {
            'time_point': time_point,
            'original_entry': deepcopy(original_entry),
            'settlement_result': deepcopy(settlement_result),
            'counterparty': counterparty,
            'timestamp': datetime.now()
        }
        if as_asset_holder:
            self.settlement_history['as_asset_holder'].append(settlement_record)
        else:
            self.settlement_history['as_liability_holder'].append(settlement_record)

class AssetLiabilityPair:
    def __init__(self,
                 time: datetime,
                 type: str,
                 amount: float,
                 denomination: str,
                 maturity_type: MaturityType,
                 maturity_date: Optional[datetime],
                 settlement_type: SettlementType,
                 settlement_denomination: str,
                 asset_holder: Agent,
                 liability_holder: Optional[Agent] = None,
                 asset_name: Optional[str] = None):  # New parameter for non-financial asset names
        self.time = time
        self.type = type
        self.amount = amount
        self.denomination = denomination
        self.maturity_type = maturity_type
        self.maturity_date = maturity_date
        self.settlement_details = SettlementDetails(
            type=settlement_type,
            denomination=settlement_denomination
        )
        self.asset_holder = asset_holder
        self.liability_holder = liability_holder
        self.asset_name = asset_name

        if type == EntryType.NON_FINANCIAL.value:
            if liability_holder is not None:
                raise ValueError("Non-financial entries cannot have a liability holder")
            if not asset_name:
                raise ValueError("Non-financial entries must have an asset name")
        else:
            if liability_holder is None:
                raise ValueError("Financial entries must have a liability holder")

    def create_entries(self) -> Tuple[BalanceSheetEntry, Optional[BalanceSheetEntry]]:
        # For delivery claim entries
        if self.type == EntryType.DELIVERY_CLAIM.value:
            if not self.asset_name:
                raise ValueError("Delivery claim must specify the asset to be delivered")

            settlement_details = SettlementDetails(
                type=SettlementType.NON_FINANCIAL_ASSET,
                denomination=self.settlement_details.denomination
            )

            # Create delivery claim (asset)
            asset_entry = BalanceSheetEntry(
                type=EntryType.DELIVERY_CLAIM,
                is_asset=True,
                counterparty=self.liability_holder.name,
                amount=self.amount,
                denomination=self.denomination,
                maturity_type=self.maturity_type,
                maturity_date=self.maturity_date,
                settlement_details=settlement_details,
                name=self.asset_name,  # Name of the asset to be delivered
                issuance_time=self.current_time_state if hasattr(self, 'current_time_state') else 't0'
            )

            # Create delivery promise (liability)
            liability_entry = BalanceSheetEntry(
                type=EntryType.DELIVERY_CLAIM,
                is_asset=False,
                counterparty=self.asset_holder.name,
                amount=self.amount,
                denomination=self.denomination,
                maturity_type=self.maturity_type,
                maturity_date=self.maturity_date,
                settlement_details=settlement_details,
                name=self.asset_name,  # Name of the asset to be delivered
                issuance_time=self.current_time_state if hasattr(self, 'current_time_state') else 't0'
            )

            return asset_entry, liability_entry

        # For payable entries (receivable-payable pairs)
        elif self.type == EntryType.PAYABLE.value:
            settlement_details = SettlementDetails(
                type=SettlementType.MEANS_OF_PAYMENT,
                denomination=self.settlement_details.denomination
            )

            # Create receivable (asset)
            asset_entry = BalanceSheetEntry(
                type=EntryType.PAYABLE,  # Both sides are PAYABLE type
                is_asset=True,
                counterparty=self.liability_holder.name if self.liability_holder else None,
                amount=self.amount,
                denomination=self.denomination,
                maturity_type=self.maturity_type,
                maturity_date=self.maturity_date,
                settlement_details=settlement_details,
                name=self.asset_name,
                issuance_time=self.current_time_state if hasattr(self, 'current_time_state') else 't0'
            )

            # Create payable (liability)
            liability_entry = BalanceSheetEntry(
                type=EntryType.PAYABLE,
                is_asset=False,
                counterparty=self.asset_holder.name,
                amount=self.amount,
                denomination=self.denomination,
                maturity_type=self.maturity_type,
                maturity_date=self.maturity_date,
                settlement_details=settlement_details,
                name=None,
                issuance_time=self.current_time_state if hasattr(self, 'current_time_state') else 't0'
            )

            return asset_entry, liability_entry

        # For non-financial entries
        if self.type == EntryType.NON_FINANCIAL.value:
            asset_entry = BalanceSheetEntry(
                type=EntryType.NON_FINANCIAL,
                is_asset=True,
                counterparty=None,
                amount=self.amount,
                denomination=self.denomination,
                maturity_type=self.maturity_type,
                maturity_date=self.maturity_date,
                settlement_details=self.settlement_details,
                name=self.asset_name,
                issuance_time='t0'  # Explicitly set issuance time
            )
            return asset_entry, None

        # For all other types (LOAN, DEPOSIT, BOND)
        asset_entry = BalanceSheetEntry(
            type=EntryType(self.type),
            is_asset=True,
            counterparty=self.liability_holder.name if self.liability_holder else None,
            amount=self.amount,
            denomination=self.denomination,
            maturity_type=self.maturity_type,
            maturity_date=self.maturity_date,
            settlement_details=self.settlement_details,
            name=self.asset_name,
            issuance_time='t0'  # Explicitly set issuance time
        )

        liability_entry = BalanceSheetEntry(
            type=EntryType(self.type),
            is_asset=False,
            counterparty=self.asset_holder.name,
            amount=self.amount,
            denomination=self.denomination,
            maturity_type=self.maturity_type,
            maturity_date=self.maturity_date,
            settlement_details=self.settlement_details,
            name=None,  # Liabilities don't need names
            issuance_time='t0'  # Explicitly set issuance time
        )

        return asset_entry, liability_entry

class EconomicSystem:
    def __init__(self):
        self.agents: Dict[str, Agent] = {}  # Current state
        self.asset_liability_pairs: List[AssetLiabilityPair] = []
        self.time_states: Dict[str, Dict[str, Agent]] = {}  # States at different time points
        self.current_time_state = "t0"  # Track current time state
        self.simulation_finalized = False  # Track if simulation is finalized
        # Initialize t0 state
        self.save_state('t0')

    def validate_time_point(self, time_point: str, allow_t0: bool = True) -> None:
        """Validate a time point string"""
        valid_points = ['t0', 't1', 't2'] if allow_t0 else ['t1', 't2']
        if time_point not in valid_points:
            raise ValueError(f"Time point must be {', '.join(valid_points)}")

    def add_agent(self, agent: Agent):
        self.agents[agent.name] = agent
        # Auto-save t0 state when agent is added
        if self.current_time_state == 't0':
            self.save_state('t0')
            
    def add_agents(self, agents: Iterable[Agent]):
        for agent in agents:
            self.add_agent(agent)

    def create_asset_liability_pair(self, pair: AssetLiabilityPair):
        self.asset_liability_pairs.append(pair)
        asset_entry, liability_entry = pair.create_entries()
        pair.asset_holder.add_asset(asset_entry)
        if liability_entry:
            pair.liability_holder.add_liability(liability_entry)

        # Auto-save state based on current time point
        self.save_state(self.current_time_state)

    def get_time_points(self) -> List[str]:
        """Get all time points in order: t0, t1, t2"""
        return ['t0', 't1', 't2']

    def save_state(self, time_point: str):
        self.validate_time_point(time_point)
        self.time_states[time_point] = {}
        for name, agent in self.agents.items():
            # 创建新Agent实例
            agent_copy = Agent(agent.name, agent.type)
            # 深拷贝所有属性和历史记录
            agent_copy.assets = deepcopy(agent.assets)
            agent_copy.liabilities = deepcopy(agent.liabilities)
            agent_copy.settlement_history = {
                'as_asset_holder': deepcopy(agent.settlement_history['as_asset_holder']),
                'as_liability_holder': deepcopy(agent.settlement_history['as_liability_holder'])
            }
            agent_copy.status = agent.status
            agent_copy.creation_time = agent.creation_time
            self.time_states[time_point][name] = agent_copy
        self.current_time_state = time_point
    def settle_entries(self, time_point: str):
        """Settle entries at a specific time point"""
        self.validate_time_point(time_point, allow_t0=False)

        # First save the current state at the previous time point
        prev_time = 't0' if time_point == 't1' else 't1'
        if prev_time not in self.time_states:
            self.save_state(prev_time)

        # Process all entries that are due at this time point
        for pair in self.asset_liability_pairs[:]:  # Create a copy to iterate over
            if (pair.maturity_type == MaturityType.FIXED_DATE and
                pair.maturity_date):
                # Check if the entry's maturity date matches our time state
                entry_time = 't1'
                if pair.maturity_date.year == 2100:  # t2
                    entry_time = 't2'

                if entry_time == time_point:
                    # Remove the original pair
                    self.asset_liability_pairs.remove(pair)
                    asset_entry, liability_entry = pair.create_entries()
                    pair.asset_holder.remove_asset(asset_entry)
                    if liability_entry:
                        pair.liability_holder.remove_liability(liability_entry)

                    # Handle settlement based on type
                    if pair.settlement_details.type == SettlementType.MEANS_OF_PAYMENT:
                        # 修改存款处理逻辑
                        remaining = pair.amount
                        # 获取所有符合条件的存款列表
                        deposits = [asset for asset in pair.liability_holder.assets if asset.type == EntryType.DEPOSIT and asset.denomination == pair.denomination]
                        if sum(d.amount for d in deposits) < remaining:
                            raise ValueError(f"Insufficient deposits for {pair.liability_holder.name}")

                        # 处理每个存款直到剩余金额为0
                        for deposit in deposits[:]:  # 使用切片创建副本以安全迭代
                            if remaining <= 0:
                                break
                            # 查找对应的银行负债
                            bank = self.agents.get(deposit.counterparty)
                            if not bank:
                                raise ValueError(f"Bank {deposit.counterparty} not found")
                            bank_liability = next(
                                (l for l in bank.liabilities
                                if l.type == EntryType.DEPOSIT
                                and l.denomination == deposit.denomination
                                and l.amount == deposit.amount
                                and l.issuance_time == deposit.issuance_time
                                and l.maturity_type == deposit.maturity_type
                                and l.counterparty == pair.liability_holder.name),  # 确保counterparty是公司A
                                None
                            )
                            
                            if not bank_liability:
                                continue  # 跳过无效存款

                            # 计算可用金额
                            amount_to_use = min(deposit.amount, remaining)
                            # 移除原存款和负债
                            pair.liability_holder.remove_asset(deposit)
                            bank.remove_liability(bank_liability)

                            # 创建结算后的存款
                            settlement_pair = AssetLiabilityPair(
                                time=datetime.now(),
                                type=EntryType.DEPOSIT.value,
                                amount=amount_to_use,
                                denomination=pair.denomination,
                                maturity_type=MaturityType.ON_DEMAND,
                                maturity_date=None,
                                settlement_type=SettlementType.NONE,
                                settlement_denomination=pair.denomination,
                                asset_holder=pair.asset_holder,
                                liability_holder=bank
                            )
                            new_asset, new_liability = settlement_pair.create_entries()
                            new_asset.issuance_time = time_point
                            new_liability.issuance_time = time_point
                            pair.asset_holder.add_asset(new_asset)
                            bank.add_liability(new_liability)
                            self.asset_liability_pairs.append(settlement_pair)

                            # 处理剩余金额
                            remaining -= amount_to_use
                            if deposit.amount > amount_to_use:
                                remainder = deposit.amount - amount_to_use
                                remainder_pair = AssetLiabilityPair(
                                    time=datetime.now(),
                                    type=EntryType.DEPOSIT.value,
                                    amount=remainder,
                                    denomination=pair.denomination,
                                    maturity_type=MaturityType.ON_DEMAND,
                                    maturity_date=None,
                                    settlement_type=SettlementType.NONE,
                                    settlement_denomination=pair.denomination,
                                    asset_holder=pair.liability_holder,
                                    liability_holder=bank
                                )
                                rem_asset, rem_liability = remainder_pair.create_entries()
                                rem_asset.issuance_time = deposit.issuance_time  # 保留原发行时间
                                rem_liability.issuance_time = deposit.issuance_time
                                pair.liability_holder.add_asset(rem_asset)
                                bank.add_liability(rem_liability)
                                self.asset_liability_pairs.append(remainder_pair)
                    elif pair.settlement_details.type == SettlementType.NON_FINANCIAL_ASSET:
                        # Find and remove the non-financial asset from the liability holder
                        non_financial_asset = next(
                            (asset for asset in pair.liability_holder.assets
                             if asset.type == EntryType.NON_FINANCIAL
                             and asset.name == pair.asset_name
                             and asset.amount >= pair.amount),
                            None
                        )

                        if not non_financial_asset:
                            raise ValueError(f"Non-financial asset {pair.asset_name} not found for settlement")

                        # Remove the asset from the liability holder
                        pair.liability_holder.remove_asset(non_financial_asset)

                        # Create non-financial asset entry for the asset holder
                        settlement_pair = AssetLiabilityPair(
                            time=datetime.now(),
                            type=EntryType.NON_FINANCIAL.value,
                            amount=pair.amount,
                            denomination=pair.settlement_details.denomination,
                            maturity_type=MaturityType.ON_DEMAND,
                            maturity_date=None,
                            settlement_type=SettlementType.NONE,
                            settlement_denomination=pair.settlement_details.denomination,
                            asset_holder=pair.asset_holder,
                            liability_holder=None,
                            asset_name=pair.asset_name
                        )
                        new_asset_entry, _ = settlement_pair.create_entries()
                        # 显式设置issuance_time
                        new_asset_entry.issuance_time = time_point


                        # Record settlement history
                        pair.asset_holder.record_settlement(
                            time_point=time_point,
                            original_entry=asset_entry,
                            settlement_result=new_asset_entry,
                            counterparty=pair.liability_holder.name,
                            as_asset_holder=True
                        )
                        pair.liability_holder.record_settlement(
                            time_point=time_point,
                            original_entry=liability_entry,
                            settlement_result=non_financial_asset,  # The non-financial asset that was delivered
                            counterparty=pair.asset_holder.name,
                            as_asset_holder=False
                        )

                        # Add entry directly to avoid default t0 issuance time
                        settlement_pair.asset_holder.add_asset(new_asset_entry)
                        self.asset_liability_pairs.append(settlement_pair)

                        # If there was remaining amount in the non-financial asset, create a new entry for it
                        if non_financial_asset.amount > pair.amount:
                            remainder_pair = AssetLiabilityPair(
                                time=datetime.now(),
                                type=EntryType.NON_FINANCIAL.value,
                                amount=non_financial_asset.amount - pair.amount,
                                denomination=non_financial_asset.denomination,
                                maturity_type=MaturityType.ON_DEMAND,
                                maturity_date=None,
                                settlement_type=SettlementType.NONE,
                                settlement_denomination=non_financial_asset.denomination,
                                asset_holder=pair.liability_holder,  # Original holder keeps remainder
                                liability_holder=None,
                                asset_name=non_financial_asset.name
                            )

                            # Create entry with current time point as issuance time
                            remainder_asset, _ = remainder_pair.create_entries()
                            remainder_asset.issuance_time = time_point

                            # Add entry
                            remainder_pair.asset_holder.add_asset(remainder_asset)
                            self.asset_liability_pairs.append(remainder_pair)

        # Auto-save state after settlements
        self.save_state(time_point)
        self.current_time_state = time_point

    def get_agents_at_time(self, time_point: str) -> Dict[str, Agent]:
        """Get agents state at a specific time point"""
        self.validate_time_point(time_point)

        # For t0, always show current state
        if time_point == 't0':
            return {name: agent for name, agent in self.agents.items()}

        # For t1 and t2, use saved state if available
        if time_point in self.time_states:
            return self.time_states[time_point]

        # If state not saved yet and we're looking at a future point,
        # we need to process settlements up to that point
        if time_point > self.current_time_state:
            # Save current state
            current_state = deepcopy(self.agents)
            current_time = self.current_time_state

            # Process settlements for each time point up to the requested one
            time_points = ['t0', 't1', 't2']
            start_idx = time_points.index(self.current_time_state) + 1
            end_idx = time_points.index(time_point) + 1

            try:
                for t in time_points[start_idx:end_idx]:
                    self.settle_entries(t)

                # Get the state after settlements
                result = {name: agent for name, agent in self.agents.items()}

                # Restore original state
                self.agents = current_state
                self.current_time_state = current_time

                return result
            except Exception as e:
                # If settlement fails, restore original state and return it
                self.agents = current_state
                self.current_time_state = current_time
                print(f"\nWarning: Could not process settlements ({str(e)})")
                return current_state

        # If none of the above, return current state
        return {name: agent for name, agent in self.agents.items()}

    def compute_changes(self, from_time: str, to_time: str) -> Dict[str, Dict[str, List]]:
        """Compute changes between two time points"""
        if from_time not in self.time_states or to_time not in self.time_states:
            raise ValueError(f"Missing state for time point {from_time} or {to_time}")

        changes = {}
        for name, to_agent in self.time_states[to_time].items():
            from_agent = self.time_states[from_time][name]

            # Find new and removed assets
            new_assets = [a for a in to_agent.assets if not any(a.matches(from_a) for from_a in from_agent.assets)]
            removed_assets = [a for a in from_agent.assets if not any(a.matches(to_a) for to_a in to_agent.assets)]

            # Find new and removed liabilities
            new_liabilities = [l for l in to_agent.liabilities if not any(l.matches(from_l) for from_l in from_agent.liabilities)]
            removed_liabilities = [l for l in from_agent.liabilities if not any(l.matches(to_l) for to_l in to_agent.liabilities)]

            changes[name] = {
                'new_assets': new_assets,
                'removed_assets': removed_assets,
                'new_liabilities': new_liabilities,
                'removed_liabilities': removed_liabilities
            }

        return changes

    def can_settle_entry(self, agent: Agent, entry: BalanceSheetEntry) -> Tuple[bool, str]:
        """Check if an agent can settle a liability"""
        if entry.settlement_details.type == SettlementType.MEANS_OF_PAYMENT:
            # Check for sufficient deposits
            deposits = sum(asset.amount for asset in agent.assets
                          if asset.type == EntryType.DEPOSIT
                          and asset.denomination == entry.denomination)
            if deposits < entry.amount:
                return False, f"Insufficient deposits: has {deposits} {entry.denomination}, needs {entry.amount}"

        elif entry.settlement_details.type == SettlementType.NON_FINANCIAL_ASSET:
            # Check for required non-financial asset
            has_asset = any(asset.type == EntryType.NON_FINANCIAL
                           and asset.name == entry.name
                           and asset.amount >= entry.amount
                           for asset in agent.assets)
            if not has_asset:
                return False, f"Missing required non-financial asset: {entry.name}"

        return True, "OK"

    def create_default_entries(self, failed_entry: BalanceSheetEntry) -> Tuple[BalanceSheetEntry, BalanceSheetEntry]:
        """Create default claim and liability entries when settlement fails"""
        # Create default claim for the creditor
        default_claim = BalanceSheetEntry(
            type=EntryType.DEFAULT,
            is_asset=True,
            counterparty=failed_entry.counterparty,
            amount=failed_entry.amount,
            denomination=failed_entry.denomination,
            maturity_type=MaturityType.ON_DEMAND,
            maturity_date=None,
            settlement_details=failed_entry.settlement_details,
            name=f"Default on {failed_entry.type.value}",
            issuance_time=self.current_time_state
        )

        # Create default liability for the debtor
        default_liability = BalanceSheetEntry(
            type=EntryType.DEFAULT,
            is_asset=False,
            counterparty=failed_entry.counterparty,
            amount=failed_entry.amount,
            denomination=failed_entry.denomination,
            maturity_type=MaturityType.ON_DEMAND,
            maturity_date=None,
            settlement_details=failed_entry.settlement_details,
            name=f"Default on {failed_entry.type.value}",
            issuance_time=self.current_time_state
        )

        return default_claim, default_liability

    def run_simulation(self) -> bool:
        """Run the full simulation from t0 through t2, handling settlements and defaults"""
        print("\nStarting simulation from t0...")

        for time_point in ['t1', 't2']:
            print(f"\nProcessing {time_point}...")

            # Get all entries that mature at this time point
            maturing_entries = []
            for agent in self.agents.values():
                for liability in agent.liabilities:
                    if (liability.maturity_type == MaturityType.FIXED_DATE and
                        ((time_point == 't1' and liability.maturity_date.year == 2050) or
                         (time_point == 't2' and liability.maturity_date.year == 2100))):
                        maturing_entries.append((agent, liability))

            # Try to settle each entry
            for agent, liability in maturing_entries:
                can_settle, reason = self.can_settle_entry(agent, liability)

                if not can_settle:
                    print(f"\nDEFAULT DETECTED: {agent.name} cannot settle {liability.type.value}")
                    print(f"Reason: {reason}")

                    # Find the corresponding asset holder
                    asset_holder = next(a for a in self.agents.values()
                                      if a.name == liability.counterparty)

                    # Remove the original asset-liability pair
                    asset_entry = next(a for a in asset_holder.assets
                                     if a.matches(liability))
                    asset_holder.remove_asset(asset_entry)
                    agent.remove_liability(liability)

                    # Create and add default entries
                    default_claim, default_liability = self.create_default_entries(liability)
                    asset_holder.add_asset(default_claim)
                    agent.add_liability(default_liability)

                    # Save state after default
                    self.save_state(time_point)
                    return False  # Stop simulation

            # If we get here, try to settle all entries for this time point
            self.settle_entries(time_point)

        print("\nSimulation completed successfully!")
        return True

    def display_settlement_history(self):
        """Display settlement history for all agents"""
        if not self.agents:
            print("\nNo agents in the system yet!")
            return

        print("\nSettlement History:")
        for agent_name, agent in self.agents.items():
            print(f"\n{agent_name}'s Settlement History:")

            # Display settlements where agent was asset holder
            if agent.settlement_history.get('as_asset_holder', []):
                print("\n  As Asset Holder:")
                for settlement in agent.settlement_history['as_asset_holder']:
                    print(f"\n    Time: {settlement['time_point']}")
                    print(f"    Original Asset: {settlement['original_entry'].type.value} "
                          f"of {settlement['original_entry'].amount} {settlement['original_entry'].denomination}")
                    print(f"    Settled For: {settlement['settlement_result'].type.value} "
                          f"of {settlement['settlement_result'].amount} {settlement['settlement_result'].denomination}")
                    print(f"    Counterparty: {settlement['counterparty']}")

            # Display settlements where agent was liability holder
            if agent.settlement_history.get('as_liability_holder', []):
                print("\n  As Liability Holder:")
                for settlement in agent.settlement_history['as_liability_holder']:
                    print(f"\n    Time: {settlement['time_point']}")
                    print(f"    Original Liability: {settlement['original_entry'].type.value} "
                          f"of {settlement['original_entry'].amount} {settlement['original_entry'].denomination}")
                    print(f"    Settled With: {settlement['settlement_result'].type.value} "
                          f"of {settlement['settlement_result'].amount} {settlement['settlement_result'].denomination}")
                    print(f"    Counterparty: {settlement['counterparty']}")

            if (not agent.settlement_history.get('as_asset_holder') and
                not agent.settlement_history.get('as_liability_holder')):
                print("  No settlements recorded")

    def display_balance_sheets(self, time_point: str):
        """Display balance sheets for all agents at a specific time point"""
        if not self.agents:
            print("\nNo agents in the system yet!")
            return

        current_agents = self.get_agents_at_time(time_point).values()
        print(f"\nBalance sheets at {time_point}:")

        for agent in current_agents:
            print(f"\n{agent.name} ({agent.type.value}):")
            print("Assets:")
            for asset in agent.assets:
                maturity_info = ""
                if asset.maturity_type == MaturityType.FIXED_DATE:
                    if asset.maturity_date.year == 2100:
                        maturity_info = " (matures at t2)"
                    elif asset.maturity_date.year == 2050:
                        maturity_info = " (matures at t1)"

                # Show appropriate entry type
                if asset.type == EntryType.PAYABLE:
                    entry_type = "receivable"
                elif asset.type == EntryType.DELIVERY_CLAIM:
                    entry_type = f"delivery claim for {asset.name}" if asset.name else "delivery claim"
                elif asset.type == EntryType.DEFAULT:
                    entry_type = f"default claim ({asset.name})"
                else:
                    entry_type = asset.type.value

                # Skip if entry has matured and been settled
                if asset.maturity_type == MaturityType.FIXED_DATE:
                    entry_maturity = 't1' if asset.maturity_date.year == 2050 else 't2'
                    if time_point > entry_maturity:
                        continue

                print(f"  - {entry_type}: {asset.amount} {asset.denomination} "
                      f"(from {asset.counterparty if asset.counterparty else 'N/A'})"
                      f"{maturity_info} [issued at {asset.issuance_time}]")

            print("Liabilities:")
            for liability in agent.liabilities:
                maturity_info = ""
                if liability.maturity_type == MaturityType.FIXED_DATE:
                    if liability.maturity_date.year == 2100:
                        maturity_info = " (matures at t2)"
                    elif liability.maturity_date.year == 2050:
                        maturity_info = " (matures at t1)"

                # Show appropriate entry type
                if liability.type == EntryType.DELIVERY_CLAIM:
                    entry_type = f"delivery promise for {liability.name}" if liability.name else "delivery promise"
                elif liability.type == EntryType.DEFAULT:
                    entry_type = f"default liability ({liability.name})"
                else:
                    entry_type = liability.type.value

                # Skip if entry has matured and been settled
                if liability.maturity_type == MaturityType.FIXED_DATE:
                    entry_maturity = 't1' if liability.maturity_date.year == 2050 else 't2'
                    if time_point > entry_maturity:
                        continue

                print(f"  - {entry_type}: {liability.amount} {liability.denomination} "
                      f"(to {liability.counterparty}){maturity_info} "
                      f"[issued at {liability.issuance_time}]")

class ExcelExporter:
    def __init__(self, system: EconomicSystem):
        self.system = system

    def create_t_table(self, sheet, row_start: int, col_start: int, agent: Agent, time_point: str):
        thick = Side(style='thick', color='000000')

        # Add time point header
        time_header = sheet.cell(row=row_start, column=1)
        time_header.value = f"Time: {time_point}"
        time_header.alignment = Alignment(horizontal="center")
        time_header.font = openpyxl.styles.Font(bold=True)

        name_cell = sheet.cell(row=row_start, column=col_start)
        name_cell.value = f"{agent.name} ({agent.type.value})"
        name_cell.alignment = Alignment(horizontal="center")

        # Set up headers and borders
        for i in range(10):
            cell = sheet.cell(row=row_start + 1, column=col_start + i)
            cell.border = Border(top=thick)

        for row in range(row_start + 1, row_start + 20):
            cell = sheet.cell(row=row, column=col_start + 4)
            cell.border = Border(right=thick)
            if row == row_start + 1:
                cell.border = Border(right=thick, top=thick)

        headers = ['Type', 'CP', 'Amount', 'Maturity', 'Settlement']
        for i, header in enumerate(headers):
            cell = sheet.cell(row=row_start + 1, column=col_start + i)
            cell.value = header
            cell.alignment = Alignment(horizontal="center")
            cell = sheet.cell(row=row_start + 1, column=col_start + i + 5)
            cell.value = header
            cell.alignment = Alignment(horizontal="center")

        # Display balance sheet entries
        current_row = row_start + 2
        for entry in agent.assets:
            # Skip entries that were issued after the current time point
            time_points = ['t0', 't1', 't2']
            if time_points.index(entry.issuance_time) > time_points.index(time_point):
                continue

            # Skip matured entries if not at t0
            if time_point != 't0' and entry.maturity_type == MaturityType.FIXED_DATE:
                entry_time = 't1' if entry.maturity_date.year == 2050 else 't2'
                if time_point > entry_time:
                    continue

            # Show entry details
            entry_type = "receivable" if entry.type == EntryType.PAYABLE else entry.type.value
            if entry.type == EntryType.NON_FINANCIAL and entry.name:
                entry_type = f"{entry.type.value} ({entry.name})"
            sheet.cell(row=current_row, column=col_start).value = entry_type
            sheet.cell(row=current_row, column=col_start + 1).value = entry.counterparty if entry.counterparty else "N/A"
            sheet.cell(row=current_row, column=col_start + 2).value = f"+{entry.amount} {entry.denomination}"

            maturity = entry.maturity_type.value
            if entry.maturity_type == MaturityType.FIXED_DATE:
                maturity = 't1' if entry.maturity_date.year == 2050 else 't2'
            sheet.cell(row=current_row, column=col_start + 3).value = f"{maturity} (issued at {entry.issuance_time})"

            settlement = entry.settlement_details.type.value
            if settlement != "none":
                settlement += f" ({entry.settlement_details.denomination})"
            sheet.cell(row=current_row, column=col_start + 4).value = settlement

            current_row += 1

        current_row = row_start + 2
        for entry in agent.liabilities:
            # Skip entries that were issued after the current time point
            time_points = ['t0', 't1', 't2']
            if time_points.index(entry.issuance_time) > time_points.index(time_point):
                continue

            # Skip matured entries if not at t0
            if time_point != 't0' and entry.maturity_type == MaturityType.FIXED_DATE:
                entry_time = 't1' if entry.maturity_date.year == 2050 else 't2'
                if time_point > entry_time:
                    continue

            # Show entry details
            entry_type = entry.type.value
            if entry.type == EntryType.DELIVERY_CLAIM:
                entry_type = f"delivery promise for {entry.name}" if entry.name else "delivery promise"
            sheet.cell(row=current_row, column=col_start + 5).value = entry_type
            sheet.cell(row=current_row, column=col_start + 6).value = entry.counterparty
            sheet.cell(row=current_row, column=col_start + 7).value = f"+{entry.amount} {entry.denomination}"

            maturity = entry.maturity_type.value
            if entry.maturity_type == MaturityType.FIXED_DATE:
                maturity = 't1' if entry.maturity_date.year == 2050 else 't2'
            sheet.cell(row=current_row, column=col_start + 8).value = f"{maturity} (issued at {entry.issuance_time})"

            settlement = entry.settlement_details.type.value
            if settlement != "none":
                settlement += f" ({entry.settlement_details.denomination})"
            sheet.cell(row=current_row, column=col_start + 9).value = settlement

            current_row += 1

        # Add totals
        total_row = current_row + 2
        sheet.cell(row=total_row, column=col_start).value = "Total Assets:"
        sheet.cell(row=total_row, column=col_start + 2).value = agent.get_total_assets()
        sheet.cell(row=total_row + 1, column=col_start).value = "Total Liabilities:"
        sheet.cell(row=total_row + 1, column=col_start + 2).value = agent.get_total_liabilities()
        sheet.cell(row=total_row + 2, column=col_start).value = "Net Worth:"
        sheet.cell(row=total_row + 2, column=col_start + 2).value = agent.get_net_worth()

        return total_row + 4
        
    def export_balance_sheets(self, output_stream: BytesIO):
        """Modified to write to in-memory buffer"""
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Balance Sheets Over Time"

        current_row = 1
        time_points = self.system.get_time_points()

        for time_point in time_points:
            agents = (self.system.time_states[time_point].values()
                     if time_point in self.system.time_states
                     else self.system.agents.values())

            sheet.cell(row=current_row, column=1).value = "=" * 50
            current_row += 1

            col_start = 2
            max_row_in_timepoint = current_row

            for agent in agents:
                max_row_in_timepoint = max(
                    max_row_in_timepoint,
                    self.create_t_table(sheet, current_row, col_start, agent, time_point)
                )
                col_start += 10

            system_total_row = max_row_in_timepoint
            sheet.cell(row=system_total_row, column=1).value = f"System Totals at {time_point}:"
            sheet.cell(row=system_total_row + 1, column=1).value = "Total Assets:"
            sheet.cell(row=system_total_row + 1, column=2).value = sum(agent.get_total_assets() for agent in agents)
            sheet.cell(row=system_total_row + 2, column=1).value = "Total Liabilities:"
            sheet.cell(row=system_total_row + 2, column=2).value = sum(agent.get_total_liabilities() for agent in agents)
            sheet.cell(row=system_total_row + 3, column=1).value = "Total Net Worth:"
            sheet.cell(row=system_total_row + 3, column=2).value = sum(agent.get_net_worth() for agent in agents)

            current_row = system_total_row + 5

        # 调整列宽
        for i in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(i)].width = 18

        wb.save(output_stream)  # 直接保存到内存流

def get_user_date_input(prompt: str) -> Optional[datetime]:
    time_state = input(prompt).strip().lower()
    if time_state == 't0':
        return datetime(2000, 1, 1)
    elif time_state == 't1':
        return datetime(2050, 1, 1)
    elif time_state == 't2':
        return datetime(2100, 1, 1)
    else:
        return None

# ======== Web应用扩展 ========
app = Flask(__name__)
app.secret_key = os.urandom(24).hex()
system = EconomicSystem()

# ======== Web界面HTML模板 ========
HTML_BASE = '''
<!DOCTYPE html>
<html>
<head>
    <title>Economic Sim Web</title>
    <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/css/bootstrap.min.css">
    <style>
        .container { 
            max-width: 800px;  
            margin: 20px auto; 
        }
        .form-section { 
            margin: 20px 0; 
            padding: 20px; 
            border: 1px solid #ddd;  
        }
        .visualization { 
            height: 400px; 
            background: #f5f5f5;  
            margin: 20px 0; 
        }
        .list-group-item form {
            margin-left: 10px;
        }
        .btn-sm {
            padding: 0.15rem 0.5rem;
            font-size: 0.875rem;
        }
    </style>
</head>
<body>
    <div class="container">
        {% with messages = get_flashed_messages(with_categories=true) %}  
            {% if messages %}
                {% for category, message in messages %}
                    <div class="alert alert-{{ category }} alert-dismissible fade show">
                        {{ message }}
                        <button type="button" class="btn-close" data-bs-dismiss="alert"></button>  
                    </div>
                {% endfor %}
            {% endif %}
        {% endwith %}
        
        <div class="d-flex justify-content-between mb-3">  
            <h2>Economic Balance Sheet Simulator</h2>
            <div>
                <a href="/export" class="btn btn-success me-2">Export to Excel</a>
                <form action="/simulate" method="post" class="d-inline"> 
                    <button class="btn btn-warning">Run Simulation</button> 
                </form>
            </div>
        </div>
        
        <div class="row">
            <div class="col-md-4">
                <div class="list-group">
                    <a href="/" class="list-group-item list-group-item-action">Home</a>
                    <a href="/batch" class="list-group-item list-group-item-action">Batch Upload</a>
                    <a href="/agents" class="list-group-item list-group-item-action">View Agents</a>
                    <a href="/settlement_history" class="list-group-item list-group-item-action">View Settlement History</a>
                </div>
            </div>
            <div class="col-md-8">
                {% block content %}{% endblock %}
            </div>
        </div>
    </div>
</body>
</html>
'''

HOME_CONTENT = '''
<div class="form-section">
    <h4>Create Agent</h4>
    <form action="/create_agent" method="post">
        <div class="mb-3">
            <input type="text" name="name" placeholder="Agent Name" class="form-control" required>
        </div>
        <select name="type" class="form-select mb-3">
            {% for type in agent_types %}
            <option value="{{ type.value }}">{{ type.value.title() }}</option>
            {% endfor %}
        </select>
        <button type="submit" class="btn btn-primary">Create</button>
    </form>
</div>

<div class="card">
    <div class="card-body">
        <h4>Create Asset-Liability Pair</h4>
        
        <form action="/create_pair" method="post" onsubmit="return validateForm()">
            <!-- Asset Holder -->
            <div class="mb-3">
                <label class="form-label">Asset Holder</label>
                <select name="asset_holder" class="form-select" required id="assetHolder">
                    {% for agent in agents %}
                    <option value="{{ agent.name }}">{{ agent.name }} ({{ agent.type.value }})</option>
                    {% endfor %}
                </select>
            </div>

            <!-- Entry Type -->
            <div class="mb-3">
                <label class="form-label">Entry Type</label>
                <select name="entry_type" class="form-select" id="entryType" required>
                    {% for et in [EntryType.LOAN, EntryType.DEPOSIT, EntryType.PAYABLE, 
                                EntryType.BOND, EntryType.DELIVERY_CLAIM, EntryType.NON_FINANCIAL] %}
                    <option value="{{ et.value }}">{{ et.value|title }}</option>
                    {% endfor %}
                </select>
            </div>

            <!-- Liability Holder (conditional) -->
            <div class="mb-3" id="liabilityHolderGroup">
                <label class="form-label">Liability Holder</label>
                <select name="liability_holder" class="form-select" id="liabilityHolder">
                    <option value="">None</option>
                    {% for agent in agents %}
                    <option value="{{ agent.name }}">{{ agent.name }}</option>
                    {% endfor %}
                </select>
            </div>

            <!-- Asset Name (conditional) -->
            <div class="mb-3" id="assetNameGroup">
                <label class="form-label">Asset Name</label>
                <input type="text" name="asset_name" class="form-control" id="assetName">
            </div>

            <!-- Amount & Denomination -->
            <div class="row mb-3">
                <div class="col">
                    <label class="form-label">Amount</label>
                    <input type="number" step="0.01" name="amount" class="form-control" required>
                </div>
                <div class="col">
                    <label class="form-label">Denomination</label>
                    <input type="text" name="denomination" class="form-control" value="USD" required>
                </div>
            </div>

            <!-- Maturity -->
            <div class="row mb-3" id="maturityGroup">
                <div class="col">
                    <label class="form-label">Maturity Type</label>
                    <select name="maturity_type" class="form-select" id="maturityType">
                        {% for mt in MaturityType %}
                        <option value="{{ mt.value }}" 
                                {% if mt == MaturityType.ON_DEMAND %}selected{% endif %}>
                            {{ mt.value|title }}
                        </option>
                        {% endfor %}
                    </select>
                </div>
                <div class="col" id="maturityDateGroup">
                    <label class="form-label">Maturity Year</label>
                    <select name="maturity_time_point" class="form-select" id="maturityDate">
                        <option value="t1">t1 (2050-01-01)</option>
                        <option value="t2">t2 (2100-01-01)</option>
                    </select>
                </div>
            </div>

            <!-- Settlement Type -->
            <div class="mb-3" id="settlementGroup">
                <label class="form-label">Settlement Type</label>
                <select name="settlement_type" class="form-select" id="settlementType">
                    {% for st in SettlementType if st != SettlementType.NONE %}
                        <option value="{{ st.value }}" 
                                {% if st == SettlementType.NONE %}selected{% endif %}>
                            {{ st.value|title }}
                        </option>
                    {% endfor %}
                </select>
            </div>

            <button type="submit" class="btn btn-primary">Create</button>
        </form>
    </div>
</div>
<script>
document.querySelector('form[action="/create_agent"]').addEventListener('submit', function(e) {
    const existingAgents = Array.from(document.querySelectorAll('#assetHolder option'))
                               .map(opt => opt.value);
    const newName = document.querySelector('input[name="name"]').value;
    
    if (existingAgents.includes(newName)) {
        e.preventDefault();
        alert('Agent name already exists!'); 
    }
});

function validateForm() {
    const entryType = document.getElementById('entryType').value;
    const assetHolder = document.getElementById('assetHolder').value;
    const liabilityHolder = document.getElementById('liabilityHolder').value;
    const maturityType = document.getElementById('maturityType');

    // 新增的验证逻辑
    if (entryType === 'non_financial') {
        if (liabilityHolder) {
            alert("Non-financial entries cannot have a liability holder!");
            return false;
        }
        if (document.getElementById('assetName').value.trim() === "") {
            alert("Non-financial entries must specify an asset name!");
            return false;
        }
    }

    // 检查资产持有者和负债持有者不同
    if (liabilityHolder && assetHolder === liabilityHolder) {
        alert("Asset holder and liability holder cannot be the same!");
        return false;
    }

    return true;
}

function filterLiabilityHolders() {
    const assetHolderSelect = document.getElementById('assetHolder');
    const liabilityHolderSelect = document.getElementById('liabilityHolder');
    const selectedAssetHolder = assetHolderSelect.value;

    // 遍历所有负债方选项
    Array.from(liabilityHolderSelect.options).forEach(option => {
        // 禁用与资产持有者相同的选项
        option.disabled = (option.value === selectedAssetHolder);
        
        // 如果是空选项则保持可用
        if (option.value === "") {
            option.disabled = false;
        }
    });
}

// 初始化时执行过滤
document.addEventListener('DOMContentLoaded', () => {
    // 给资产持有者下拉菜单添加事件监听
    const assetHolderSelect = document.getElementById('assetHolder');
    assetHolderSelect.addEventListener('change', filterLiabilityHolders);
    
    // 初始执行一次过滤
    filterLiabilityHolders();
    
    // 保持原有的表单更新逻辑
    updateFormFields();
    document.getElementById('entryType').addEventListener('change', updateFormFields);
});

// 修改原有表单验证函数，增加对同一主体的检查
function validateForm() {
    const assetHolder = document.getElementById('assetHolder').value;
    const liabilityHolder = document.getElementById('liabilityHolder').value;
    
    // 新增验证：资产持有者和负债持有者不能相同
    if (liabilityHolder && assetHolder === liabilityHolder) {
        alert('Asset Holder and Liability Holder cannot be the same!');
        return false;
    }
    
    // 保持原有的其他验证逻辑
    const entryType = document.getElementById('entryType').value;
    // ... [原有其他验证逻辑保持不变] ...
    
    return true;
}


function updateFormFields() {
    const entryType = document.getElementById('entryType').value;
    const formElements = {
        liability: {
            group: document.getElementById('liabilityHolderGroup'),
            field: document.getElementById('liabilityHolder')
        },
        assetName: {
            group: document.getElementById('assetNameGroup'),
            field: document.getElementById('assetName')
        },
        maturity: {
            group: document.getElementById('maturityGroup'),
            type: document.getElementById('maturityType'),
            date: document.getElementById('maturityDate')
        },
        settlement: {
            group: document.getElementById('settlementGroup'),
            field: document.getElementById('settlementType')
        }
    };

    // 重置所有字段状态
    formElements.liability.group.style.display = 'block';
    formElements.liability.field.disabled = false;
    formElements.liability.field.required = true;
    formElements.assetName.group.style.display = 'none';
    formElements.assetName.field.required = false;
    formElements.maturity.group.style.display = 'block';
    formElements.maturity.type.disabled = false;
    formElements.maturity.date.disabled = false;
    formElements.settlement.group.style.display = 'block';
    formElements.settlement.field.disabled = false;

    // 处理 non_financial 类型
    if (entryType === 'non_financial') {

        document.getElementById('settlementType').value = 'none';
        document.getElementById('settlementType').disabled = true;
        document.getElementById('settlementGroup').style.display = 'none';
        document.getElementById('maturityType').value = 'on_demand';
        document.getElementById('liabilityHolder').value = '';

        // 强制设置字段值
        formElements.settlement.field.value = 'none';
        formElements.maturity.type.value = 'on_demand';
        formElements.liability.field.value = '';

        // 隐藏并禁用字段组
        formElements.liability.group.style.display = 'none';
        formElements.maturity.group.style.display = 'none';
        formElements.settlement.group.style.display = 'none';

        // 禁用字段输入
        formElements.liability.field.disabled = true;
        formElements.liability.field.required = false;
        formElements.maturity.type.disabled = true;
        formElements.maturity.date.disabled = true;
        formElements.settlement.field.disabled = true;

        // 强制显示资产名称字段
        formElements.assetName.group.style.display = 'block';
        formElements.assetName.field.required = true;
    } else if (entryType === 'delivery_claim') {
        formElements.assetName.group.style.display = 'block';
        formElements.assetName.field.required = true;
        formElements.settlement.field.value = 'non_financial_asset';
        formElements.settlement.field.disabled = true;
    } else if (['loan', 'bond'].includes(entryType)) {
        formElements.maturity.group.style.display = 'block';
        formElements.maturity.type.disabled = false;
        formElements.settlement.group.style.display = 'block';
    } else {
        formElements.maturity.group.style.display = 'none';
        formElements.maturity.type.disabled = true;
        formElements.maturity.date.disabled = true;
        formElements.settlement.group.style.display = 'none';
        formElements.settlement.field.disabled = true;
    }

    // 处理 payable 类型
    if (entryType === 'payable') {
        document.getElementById('assetName').value = '';
        formElements.assetName.group.style.display = 'none';
        formElements.assetName.field.required = false;
        formElements.maturity.group.style.display = 'block'; 
        formElements.maturity.type.disabled = false; // 允许选择到期类型
        formElements.maturity.date.disabled = false; // 允许选择到期时间点
        formElements.settlement.group.style.display = 'none'; 
        formElements.settlement.field.disabled = true;
        formElements.settlement.field.value = 'means_of_payment'; 
    }
    const maturityType = document.getElementById('maturityType').value;
    const maturityDateGroup = document.getElementById('maturityDateGroup');

    // 仅当maturityGroup可见时处理
    if (document.getElementById('maturityGroup').style.display !== 'none') {
        maturityDateGroup.style.display = maturityType === 'fixed_date' ? 'block' : 'none';
    } else {
        maturityDateGroup.style.display = 'none';
    }
}

document.addEventListener('DOMContentLoaded', () => {
    updateFormFields();
    document.getElementById('entryType').addEventListener('change', updateFormFields);
    document.getElementById('maturityType').addEventListener('change', updateFormFields);
});


function updateFormGuidance(entryType, { liability, assetName }) {
    // 清理旧指引
    ['deliveryClaimAlert', 'nonFinancialDesc'].forEach(id => {
        const el = document.getElementById(id);
        el && el.remove();
    });

    // 交割请求权指引
    if (entryType === 'delivery_claim') {
        const guidance = document.createElement('div');
        guidance.id = 'deliveryClaimAlert';
        guidance.className = 'alert alert-warning mt-3';
        guidance.innerHTML = `
            <i class="bi bi-exclamation-triangle me-2"></i>
            Required: Liability holder must possess matching non-financial assets
        `;
        liability.group.after(guidance);
    }

    // 非金融资产描述
    if (entryType === 'non_financial') {
        const desc = document.createElement('small');
        desc.id = 'nonFinancialDesc';
        desc.className = 'form-text text-muted mt-1';
        desc.textContent = 'Examples: Physical assets (equipment, inventory) or intellectual property';
        assetName.field.after(desc);
    }
}

// 移除重复的initForm定义,整合初始化逻辑
function initForm() {
    updateFormFields(); 
    // 添加指引初始化
    const entryType = document.getElementById('entryType').value;
    updateFormGuidance(entryType, {
        liability: {
            group: document.getElementById('liabilityHolderGroup'),
            field: document.getElementById('liabilityHolder')
        },
        assetName: {
            group: document.getElementById('assetNameGroup'),
            field: document.getElementById('assetName')
        }
    });
}

document.addEventListener('DOMContentLoaded', () => {
    filterLiabilityHolders();
    updateFormFields();
    initForm();
    
    document.getElementById('assetHolder').addEventListener('change', filterLiabilityHolders);
    document.getElementById('entryType').addEventListener('change', () => {
        updateFormFields();
        updateFormGuidance(document.getElementById('entryType').value, {
            liability: {
                group: document.getElementById('liabilityHolderGroup'),
                field: document.getElementById('liabilityHolder')
            },
            assetName: {
                group: document.getElementById('assetNameGroup'),
                field: document.getElementById('assetName')
            }
        });
    });
    document.getElementById('maturityType').addEventListener('change', updateFormFields);
});
</script>
'''

BATCH_CONTENT = '''
<div class="form-section">
    <h4>Batch Upload</h4>
    <form action="/batch_upload" method="post" enctype="multipart/form-data">
        <div class="mb-3">
            <label class="form-label">Upload CSV/Excel:</label>
            <input type="file" name="file" class="form-control" accept=".csv,.xlsx" required>
        </div>
        <div class="alert alert-info">
            <small>File format: Agent1,Agent2,Amount,Type,Denomination</small>
        </div>
        <button type="submit" class="btn btn-primary">Upload</button>
    </form>
</div>
'''

# ======== Web路由处理 ========
@app.route('/')
def home():
    return render_template_string(
        HTML_BASE + HOME_CONTENT,
        agent_types=AgentType,
        agents=system.agents.values(),
        entry_types=[EntryType.LOAN, EntryType.DEPOSIT, EntryType.PAYABLE, 
                     EntryType.DELIVERY_CLAIM, EntryType.NON_FINANCIAL],
        MaturityType=MaturityType,           # 改成大写，匹配模板中的变量
        SettlementType=SettlementType,       # 同上
        EntryType=EntryType
    )


@app.route('/create_agent', methods=['POST'])
def create_agent():
    name = request.form['name']
    agent_type = AgentType(request.form['type'])
    
    if name in system.agents:
        flash(f"Agent '{name}' already exists!", "danger")  # 使用flash提示替代直接返回
        return redirect(url_for('home'))  # 保持在当前页面
    
    agent = Agent(name, agent_type)
    system.add_agent(agent)
    flash(f"Agent '{name}' created successfully!", "success")
    return redirect(url_for('home'))

@app.route('/create_pair', methods=['POST'])
def create_pair():
    try:
        # 基础字段验证
        asset_holder_name = request.form['asset_holder']
        entry_type = EntryType(request.form['entry_type'])
        amount = float(request.form['amount'])
        denomination = request.form.get('denomination', 'USD')

        # 获取Agent实例
        asset_holder = system.agents.get(asset_holder_name)
        if not asset_holder:
            raise ValueError(f"Asset holder {asset_holder_name} not found")

        # 初始化参数
        liability_holder = None
        asset_name = None
        settlement_type = SettlementType.NONE
        maturity_type = MaturityType.ON_DEMAND
        maturity_date = None

        # 处理非金融资产的特殊验证
        if entry_type == EntryType.NON_FINANCIAL:
            # 强制资产名称检查
            asset_name = request.form.get('asset_name')
            settlement_type = SettlementType.NONE  # 强制设置为NONE
            maturity_type = MaturityType.ON_DEMAND  # 强制设置为ON_DEMAND
            liability_holder = None
            if not asset_name:
                raise ValueError("Non-financial entries must have an asset name")
            



        # 处理金融条目
        else:
            # 负债方必须存在
            liability_holder_name = request.form.get('liability_holder')
            if not liability_holder_name:
                raise ValueError("Liability holder is required for financial entries")
            
            liability_holder = system.agents.get(liability_holder_name)
            if not liability_holder:
                raise ValueError(f"Liability holder {liability_holder_name} not found")
            
            # 禁止自交易
            if liability_holder == asset_holder:
                raise ValueError("Asset holder and liability holder cannot be the same")

            # 处理交付索赔的特殊验证
            if entry_type == EntryType.DELIVERY_CLAIM:
                asset_name = request.form.get('asset_name')
                if not asset_name:
                    raise ValueError("Delivery claim requires asset name")
                settlement_type = SettlementType.NON_FINANCIAL_ASSET
            if entry_type == EntryType.PAYABLE:
                # 应付账款必须使用支付手段结算
                settlement_type = SettlementType.MEANS_OF_PAYMENT
                required_amount = amount
                available_funds = sum(
                    a.amount for a in liability_holder.assets
                    if a.type == EntryType.DEPOSIT  # 检查存款类资产
                    and a.denomination == denomination  # 币种匹配
                )
                if available_funds < required_amount:
                    raise ValueError(
                        f"{liability_holder.name} has insufficient funds: "
                        f"{available_funds} {denomination} available, needs {required_amount}"
                    )

                


            # 处理到期日期
            maturity_type = MaturityType(request.form.get('maturity_type', MaturityType.ON_DEMAND.value))
            if maturity_type == MaturityType.FIXED_DATE:
                maturity_time_point = request.form.get('maturity_time_point', 't1')
                maturity_date = (
                    datetime(2050, 1, 1) if maturity_time_point == 't1'
                    else datetime(2100, 1, 1)
                )

            # 处理结算类型
            if entry_type in (EntryType.LOAN, EntryType.BOND):
                settlement_type = SettlementType(
                    request.form.get('settlement_type', SettlementType.NONE.value)
                )

        # 创建资产-负债对
        pair = AssetLiabilityPair(
            time=datetime.now(),
            type=entry_type.value,
            amount=amount,
            denomination=denomination,
            maturity_type=maturity_type,
            maturity_date=maturity_date,
            settlement_type=settlement_type,
            settlement_denomination=denomination,
            asset_holder=asset_holder,
            liability_holder=liability_holder,
            asset_name=asset_name
        )

        system.create_asset_liability_pair(pair)
        flash("Asset-liability pair created successfully!", "success")
        return redirect(url_for('home'))

    except (KeyError, ValueError) as e:  # 合并预期内的异常
        error_msg = str(e)
        if isinstance(e, KeyError):
            error_msg = f"Missing required field: {e}"
        flash(error_msg, "danger")
        app.logger.warning(f"Validation error: {error_msg}")
        return redirect(url_for('home'))  # 统一重定向到主页

    except Exception as e:
        app.logger.exception("Unexpected error in create_pair")  # 记录完整堆栈
        flash("Internal server error. Please contact support.", "danger")
        return redirect(url_for('home'))



@app.route('/batch')
def batch_page():
    return render_template_string(HTML_BASE + BATCH_CONTENT)

@app.route('/batch_upload', methods=['POST'])
def batch_upload():
    if 'file' not in request.files:
        return "No file uploaded", 400
    
    file = request.files['file']
    if file.filename.endswith('.csv'):
        df = pd.read_csv(file)
    elif file.filename.endswith('.xlsx'):
        df = pd.read_excel(file)
    else:
        return "Unsupported file type", 400
    
    for _, row in df.iterrows():
        try:
            # 自动创建不存在的Agent
            for agent_name in [row[0], row[1]]:
                if agent_name not in system.agents:
                    system.add_agent(Agent(agent_name, AgentType.OTHER))
            
            asset_holder = system.agents[row[0]]
            liability_holder = system.agents[row[1]] if pd.notna(row[1]) else None
            
            pair = AssetLiabilityPair(
                time=datetime.now(),
                type=row[3] if len(row) >3 else EntryType.LOAN.value,
                amount=float(row[2]),
                denomination=row[4] if len(row) >4 else "USD",
                maturity_type=MaturityType.ON_DEMAND,
                maturity_date=None,
                settlement_type=SettlementType.MEANS_OF_PAYMENT,
                settlement_denomination="USD",
                asset_holder=asset_holder,
                liability_holder=liability_holder
            )
            system.create_asset_liability_pair(pair)
        except Exception as e:
            continue  # 简化错误处理
    
    return redirect('/')

@app.route('/agents')
def list_agents():
    agents_html = "<ul class='list-group'>"
    for agent in system.agents.values():
        agents_html += f'''
        <li class="list-group-item">
            <h5><a href="/agent/{agent.name}" class="text-decoration-none">{agent.name}</a> ({agent.type.value})</h5>
            <div class="row">
                <div class="col">
                    <h6>Assets</h6>
                    <ul class="list-group">
                        {"".join(
                            f'<li class="list-group-item">'
                            f'{a.name + ": " if a.type == EntryType.NON_FINANCIAL else ""}'
                            f'{a.amount} {a.denomination} ({a.type.value.title()}) '
                            f'[issued at {a.issuance_time}]</li>'
                            for a in agent.assets
                        )}
                    </ul>
                </div>
                <div class="col">
                    <h6>Liabilities</h6>
                    <ul class="list-group">
                        {"".join(
                            f'<li class="list-group-item">'
                            f'{l.amount} {l.denomination} ({l.type.value.title()}) '
                            f'[issued at {l.issuance_time}]</li>'
                            for l in agent.liabilities
                        )}
                    </ul>
                </div>
            </div>
        </li>
        '''
    agents_html += "</ul>"
    
    return render_template_string(HTML_BASE + '''
        <div class="form-section">
            <h4>All Agents</h4>
            ''' + agents_html + '''
        </div>
    ''')

@app.route('/agent/<name>')
def agent_detail(name):
    agent = system.agents.get(name)
    if not agent:
        return "Agent not found", 404
    
    # 修改这里：使用针对具体代理的session键
    session_key = f'edit_mode_{name}'
    edit_mode = session.get(session_key, False)  # 获取该代理的编辑状态
    
    return render_template_string(
        AGENT_TEMPLATE,
        agent=agent,
        edit_mode=edit_mode,
        EntryType=EntryType
    )

AGENT_TEMPLATE = HTML_BASE + '''
<div class="card mb-3">
    <div class="card-header d-flex justify-content-between align-items-center"> 
        <h5>
        <a href="{{ url_for('agent_detail', name=agent.name) }}" class="text-decoration-none">
            {{ agent.name }}
        </a> 
        ({{ agent.type.value }})
        </h5>
        <div>
            <!-- Edit Mode Toggle Button -->
            <a href="{{ url_for('toggle_edit_mode', name=agent.name) }}" 
               class="btn btn-sm {{ 'btn-warning' if edit_mode else 'btn-outline-secondary' }}">
                {{ 'Exit Edit' if edit_mode else 'Edit' }}
            </a>
            <!-- Delete Agent Button (only shown when no assets/liabilities) -->
            {% if edit_mode and agent.assets|length == 0 and agent.liabilities|length == 0 %}
            <form action="{{ url_for('delete_agent', name=agent.name) }}" 
                  method="post" 
                  class="d-inline ms-2">
                <button type="submit" 
                        class="btn btn-sm btn-danger"
                        onclick="return confirm('Are you sure you want to delete this agent?')"> 
                    Delete Agent
                </button>
            </form>
            {% endif %}
        </div>
    </div>

    <div class="card-body">
        <div class="row">
            <!-- Assets Column -->
            <div class="col-md-6">
                <h6 class="mb-3">Assets</h6>
                <ul class="list-group">
                    {% for asset in agent.assets %}
                    <li class="list-group-item d-flex justify-content-between align-items-center">
                        <div class="me-3">
                            <span class="badge bg-primary me-2">{{ asset.type.value|upper }}</span>
                            {{ asset.amount }} {{ asset.denomination }}
                            {% if asset.counterparty %}
                            <div class="text-muted small mt-1">From {{ asset.counterparty }}</div>
                            {% endif %}
                        </div>
                        {% if edit_mode %}
                        <form action="{{ url_for('delete_entry') }}" 
                              method="post"
                              onsubmit="return confirm('Are you sure you want to delete this asset?')">
                            <input type="hidden" name="entry_type" value="asset">
                            <input type="hidden" name="agent_name" value="{{ agent.name }}">
                            <input type="hidden" name="amount" value="{{ asset.amount }}">
                            <input type="hidden" name="denomination" value="{{ asset.denomination }}">
                            <input type="hidden" name="counterparty" value="{{ asset.counterparty or '' }}">
                            <input type="hidden" name="entry_type_value" value="{{ asset.type.value }}">
                            <button type="submit" 
                                    class="btn btn-danger btn-sm"
                                    title="Delete Asset">
                                &times;
                            </button>
                        </form>
                        {% endif %}
                    </li>
                    {% endfor %}
                </ul>
            </div>

            <!-- Liabilities Column -->
            <div class="col-md-6">
                <h6 class="mb-3">Liabilities</h6>
                <ul class="list-group">
                    {% for liability in agent.liabilities %}
                    <li class="list-group-item d-flex justify-content-between align-items-center">
                        <div class="me-3">
                            <span class="badge bg-danger me-2">{{ liability.type.value|upper }}</span>
                            {{ liability.amount }} {{ liability.denomination }}
                            <div class="text-muted small mt-1">From {{ liability.counterparty }}</div>
                        </div>
                        {% if edit_mode %}
                        <form action="{{ url_for('delete_entry') }}" 
                              method="post"
                              onsubmit="return confirm('Are you sure you want to delete this liability?')">
                            <input type="hidden" name="entry_type" value="liability">
                            <input type="hidden" name="agent_name" value="{{ agent.name }}">
                            <input type="hidden" name="amount" value="{{ liability.amount }}">
                            <input type="hidden" name="denomination" value="{{ liability.denomination }}">
                            <input type="hidden" name="counterparty" value="{{ liability.counterparty or '' }}">
                            <input type="hidden" name="entry_type_value" value="{{ liability.type.value }}">
                            <button type="submit" 
                                    class="btn btn-danger btn-sm"
                                    title="Delete Liability">
                                &times;
                            </button>
                        </form>
                        {% endif %}
                    </li>
                    {% endfor %}
                </ul>
            </div>
        </div>
    </div>
</div>
'''


@app.route('/settlement_history')
def settlement_history():
    agents = system.agents.values()
    return render_template_string(
        HTML_BASE + '''
        <div class="card">
            <div class="card-header">
                <h5>All Agents' Settlement History</h5>
            </div>
            <div class="card-body">
                {% for agent in agents %}
                <div class="mb-4">
                    <h6>{{ agent.name }} ({{ agent.type.value }})</h6>
                    <div class="ms-4">
                        <h7>As Asset Holder:</h7>
                        {% for record in agent.settlement_history['as_asset_holder'] %}
                        <div class="alert alert-secondary mt-2">
                            <p class="mb-1"><strong>Time:</strong> {{ record.time_point }}</p>
                            <p class="mb-1"><strong>Type:</strong> {{ record.original_entry.type.value|title }}</p>
                            <p class="mb-1"><strong>Amount:</strong> {{ record.original_entry.amount }} {{ record.original_entry.denomination }}</p>
                            <p class="mb-0"><strong>Settled:</strong> {{ record.settlement_result.type.value|title }} ({{ record.settlement_result.amount }} {{ record.settlement_result.denomination }})</p>
                        </div>
                        {% else %}
                        <div class="alert alert-light">No asset holder settlements</div>
                        {% endfor %}

                        <h7 class="mt-3">As Liability Holder:</h7>
                        {% for record in agent.settlement_history['as_liability_holder'] %}
                        <div class="alert alert-warning mt-2">
                            <p class="mb-1"><strong>Time:</strong> {{ record.time_point }}</p>
                            <p class="mb-1"><strong>Type:</strong> {{ record.original_entry.type.value|title }}</p>
                            <p class="mb-1"><strong>Amount:</strong> {{ record.original_entry.amount }} {{ record.original_entry.denomination }}</p>
                            <p class="mb-0"><strong>Settled:</strong> {{ record.settlement_result.type.value|title }} ({{ record.settlement_result.amount }} {{ record.settlement_result.denomination }})</p>
                        </div>
                        {% else %}
                        <div class="alert alert-light">No liability holder settlements</div>
                        {% endfor %}
                    </div>
                </div>
                <hr>
                {% endfor %}
            </div>
        </div>
        ''',
        agents=agents
    )

@app.route('/simulate', methods=['POST'])
def run_simulation():
    try:
        success = system.run_simulation()
        message = "Simulation completed successfully!" if success else "Simulation failed due to payment defaults"
        return render_template_string(SIMULATION_RESULT, message=message)
    except Exception as e:
        return f"Simulation error: {str(e)}", 500

SIMULATION_RESULT = '''
<div class="alert alert-info">
    <h4>Simulation Result</h4>
    <p>{{ message }}</p>
    <a href="/" class="btn btn-secondary">Back to Main</a>
</div>
'''

@app.route('/export')
def export_data():
    if not EXCEL_AVAILABLE:
        return "Excel support not enabled. Install openpyxl first.", 503
    
    try:
        # 创建内存文件流
        output = BytesIO()
        
        # 使用现有的Excel导出逻辑
        exporter = ExcelExporter(system)
        
        # 生成工作簿到内存流
        exporter.export_balance_sheets(output)  # 需要修改原ExcelExporter使其支持流输出
        
        # 重置指针位置
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name="economic_simulation.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        app.logger.error(f"Export failed: {str(e)}")
        return f"Export failed: {str(e)}", 500

@app.context_processor
def inject_enums():
    return {
        'agent_types': AgentType,
        'entry_types': [et for et in EntryType if et != EntryType.DEFAULT],
        'settlement_types': SettlementType,
        'maturity_types': MaturityType
    }

@app.route('/toggle_edit_mode/<name>')
def toggle_edit_mode(name):
    session_key = f'edit_mode_{name}'
    # 切换该代理的编辑状态
    session[session_key] = not session.get(session_key, False) 
    return redirect(url_for('agent_detail', name=name))

# 删除条目
@app.route('/delete_entry', methods=['POST'])
def delete_entry():
    try:
        # 获取所有必要参数
        agent_name = request.form['agent_name']
        entry_type = request.form['entry_type']  # asset/liability
        amount = float(request.form['amount'])
        denomination = request.form['denomination']
        counterparty = request.form.get('counterparty', '').strip() or None  # 处理空字符串
        entry_type_value = EntryType(request.form['entry_type_value'])
        
        # 获取Agent实例
        agent = system.agents.get(agent_name)
        if not agent:
            flash("Agent not found", "danger")
            return redirect(url_for('home'))

        target_entry = None
        if entry_type == 'asset':
            # 在资产中查找匹配条目
            target_entry = next((a for a in agent.assets 
                               if a.amount == amount
                               and a.denomination == denomination
                               and a.counterparty == counterparty
                               and a.type == entry_type_value), None)
            if target_entry:
                # 仅处理金融资产的对应负债
                if entry_type_value != EntryType.NON_FINANCIAL and counterparty:
                    counterpart = system.agents.get(counterparty)
                    if counterpart:
                        # 查找并删除对手方的负债
                        counterpart_liability = next((l for l in counterpart.liabilities 
                                                     if l.amount == amount
                                                     and l.denomination == denomination
                                                     and l.counterparty == agent_name
                                                     and l.type == entry_type_value), None)
                        if counterpart_liability:
                            counterpart.remove_liability(counterpart_liability)
                # 删除资产条目
                agent.remove_asset(target_entry)
                flash("Asset deleted successfully", "success")
        elif entry_type == 'liability':
            # 在负债中查找匹配条目
            target_entry = next((l for l in agent.liabilities 
                               if l.amount == amount
                               and l.denomination == denomination
                               and l.counterparty == counterparty
                               and l.type == entry_type_value), None)
            if target_entry:
                # 仅处理金融负债的对应资产
                if entry_type_value != EntryType.NON_FINANCIAL and counterparty:
                    counterpart = system.agents.get(counterparty)
                    if counterpart:
                        # 查找并删除对手方的资产
                        counterpart_asset = next((a for a in counterpart.assets 
                                                if a.amount == amount
                                                and a.denomination == denomination
                                                and a.counterparty == agent_name
                                                and a.type == entry_type_value), None)
                        if counterpart_asset:
                            counterpart.remove_asset(counterpart_asset)
                # 删除负债条目
                agent.remove_liability(target_entry)
                flash("Liability deleted successfully", "success")
        
        if target_entry:
            # 从资产-负债对中移除对应的条目
            system.asset_liability_pairs = [
                pair for pair in system.asset_liability_pairs
                if not (
                    (pair.asset_holder.name == agent_name and
                     any(a.matches(target_entry) for a in pair.asset_holder.assets)) or
                    (pair.liability_holder and 
                     pair.liability_holder.name == agent_name and
                     any(l.matches(target_entry) for l in pair.liability_holder.liabilities))
                )
            ]
        else:
            flash("Entry not found", "danger")
        
        return redirect(url_for('agent_detail', name=agent_name))
    
    except Exception as e:
        flash(f"Error deleting entry: {str(e)}", "danger")
        return redirect(url_for('home'))

# 删除Agent
@app.route('/delete_agent/<name>', methods=['POST'])
def delete_agent(name):
    agent = system.agents.get(name)
    if not agent:
        flash("Agent not found", "danger")
        return redirect(url_for('home'))
    
    if len(agent.assets) > 0 or len(agent.liabilities) > 0:
        flash("Cannot delete agent with existing assets/liabilities", "danger")
        return redirect(url_for('agent_detail', name=name))
    
    # 从所有时间点状态中删除
    for time_point in system.time_states:
        if name in system.time_states[time_point]:
            del system.time_states[time_point][name]
    
    # 从当前Agent列表中删除
    del system.agents[name]
    flash("Agent deleted successfully", "success")
    return redirect(url_for('home'))

# ======== 运行应用 ========
if __name__ == '__main__':
    app.run(debug=True, port=5000)